import streamlit as st
import calendar
from datetime import date, datetime
import pandas as pd
from common import gradient_line, create_workbook, header_settings, subheader_settings, sheet_title_settings
from tools.pivot import pivot_sheet
from tools.share_of_voice import share_of_voice
from tools.media_statistics import media_statistics
from tools.daily_statistics import daily_statistics
from tools.tonality_sheets import tonality_sheets
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook

import re

st.cache_data()
def get_categories(df):
    return sorted(list(set(df['Category'].to_list())))

def basic_report():
    
    col11, col12 = st.columns([1,2])
    with col11:
        st.header('🗄️File Handler')
        gradient_line()
        with st.container(border=True):
            raw_file = st.file_uploader(
                label='Upload Cleaned File',
                type=['xls', 'xlsx'])
                        
            if raw_file:

                df = pd.read_excel(raw_file)
                df['Ad Value'] = df['Ad Value'].astype(str)
                df['Ad Value'] = df['Ad Value'].str.replace(',', '')
                df['Ad Value'] = df['Ad Value'].astype(float)

                initial_categories = get_categories(df)

                with col11:
                    with st.container(border=True):
                        out_fname = st.text_input(label='Output Filename')
                
                if out_fname:
                    with col12:
                        st.header('⚙️Parameters')
                        gradient_line()
                        cola1, cola2 = st.columns([2,3], border=True)
                        with cola1:
                            st.subheader('Report Information')
                            gradient_line()
                            
                            with st.container(border=True):
                                report_client = st.text_input(
                                    label='Client Name')
                        
                            with st.container(border=True):
                                date_selector = st.radio(
                                    label='Date Selector',
                                    options=['Date', 'Date Range'])

                                if date_selector=='Date':
                                    cold1, cold2 = st.columns(2)
                                    with cold1:
                                        month_select = st.selectbox(
                                            label='Month',
                                            options=list(calendar.month_name)[1:])
                                    with cold2:
                                        now_year = date.today().year
                                        year_select = st.selectbox(
                                            label='Year',
                                            options=[year for year in range(now_year, now_year - 6, -1)])
                                    
                                    _date = f'{month_select} {year_select}'
                                    sub_header = f'Media Meter Inc. Report for {_date}'

                                else:
                                    start_date = datetime.now().date()
                                    end_date = datetime.now().date()
                                    
                                    date_range = st.date_input(
                                        label="Select date range",
                                        value=(start_date, end_date),
                                        min_value=date(2000, 1, 1),
                                        max_value=date.today())
                                    
                                    try:
                                        m1 = date_range[0].strftime('%B')
                                        d1 = date_range[0].day
                                        y1 = date_range[0].year
                                        m2 = date_range[1].strftime('%B')
                                        d2 = date_range[1].day
                                        y2 = date_range[1].year
                                    except:
                                        pass
                                    else:
                                        if y1==y2:
                                            _date = f'{m1} {d1} - {m2} {d2}, {y2}'
                                        else:
                                            _date = f'{m1} {d1}, {y1} - {m2} {d2,} {y2}'
                                        
                                        sub_header = f'Media Meter Inc. Report for {_date}'


                        with cola2:
                            st.subheader('Category Select')
                            gradient_line()
                            cola21, cola22, cola23 = st.columns(3, border=True)
                            with cola21:
                                main_category_select = st.multiselect(
                                    label='Main Category',
                                    options=initial_categories)

                                if main_category_select != []:
                                    for category in main_category_select:
                                        initial_categories.remove(category)
                        
                            with cola22:
                                competitor_category_select = st.multiselect(
                                    label='Competitor Category',
                                    options=initial_categories)

                                if competitor_category_select != []:
                                    for category in competitor_category_select:
                                        initial_categories.remove(category)
                        
                            with cola23:
                                industry_category_select = st.multiselect(
                                    label='Industry Category',
                                    options=initial_categories)
                        
                            colb1, colb2 = st.columns([2, 1], border=True)
                            with colb1:
                                st.subheader('Sheet Layout')
                                gradient_line()
                                colb11, colb12 = st.columns(2, border=True)
                                with colb11:
                                    main_sheet_option = st.radio(
                                        label='Main Sheet',
                                        options=['Single', 'Multiple'])

                                with colb12:
                                    competitor_sheet_option = st.radio(
                                        label='Competitor Sheet',
                                        options=['Single', 'Multiple'])
                                    
                                st.subheader('Additional Sheets')
                                gradient_line()
                                with st.container(border=True):
                                    cb_tonality_sheets = st.checkbox(label='Tonality Sheets')

                                # st.subheader('Sheet Settings')
                                gradient_line()
                                with st.popover('Sheet Settings', width='stretch'):
                                    tab1, tab2, tab3 = st.tabs(['Header', 'Sub-Header', 'Title'])
                                    with tab1:
                                        header_font_name, header_font_size, header_font_color, header_bold, header_italic = header_settings()
                                    with tab2:
                                        subheader_font_name, subheader_font_size, subheader_font_color, subheader_bold, subheader_italic = subheader_settings()
                                    with tab3:
                                        sheettitle_font_name, sheettitle_font_size, sheettitle_font_color, sheettitle_bold, sheettitle_italic = sheet_title_settings()
                                    

                                    
                            
                            with colb2:
                                st.subheader('Chart Select')
                                gradient_line()
                                with st.container(border=True):
                                    daily_stats = st.checkbox(
                                        label='Daily Statistics')
                                    media_stats = st.checkbox(
                                        label='Media Statistics')
                                    share_voice = st.checkbox(
                                        label='Share of Voice')
                        
                        
                                
                        
                        if report_client and (main_category_select or competitor_category_select):
                            btn_create_br = st.button(
                                label='Create Report')

                            if btn_create_br:

                                # create a blank excel file
                                create_workbook(out_fname)                                
                                
                                report_name = f'{out_fname}.xlsx'
                                writer = pd.ExcelWriter(report_name, engine='openpyxl', mode='a', if_sheet_exists='overlay')

                                # Basic Sheets
                                categories = {}
                                if main_category_select != []:
                                    categories['Company'] = [main_category_select, main_sheet_option]
                                if competitor_category_select != []:
                                    categories['Competitors'] = [competitor_category_select, competitor_sheet_option]
                                if industry_category_select != []:
                                    categories['Industry'] = [industry_category_select, 'Single']
                                
                                for category_name, _data in categories.items():
                                    category = _data[0]
                                    options = _data[1]

                                    category_df = pivot_sheet(
                                        data_frame=df,
                                        category=category)
                                    
                                    if options=='Single':
                                        initial_data = []
                                        for company, data_frame in category_df.items():
                                            initial_data.append(data_frame)
                                        
                                        final_data = pd.concat(initial_data)
                                        final_data = final_data.rename(columns={'Bucket':'Company', 'Raw Date':'Date', 'AVE':'PR Value', 'Article ID':'Count'})
                                        final_data.to_excel(writer, sheet_name=category_name, startrow=6, startcol=0, index=False)

                                        ws = writer.sheets[category_name]
                                        ws.cell(row=5, column=1).value = f'{category_name} News'
                                        ws.cell(row=2, column=1).value = sub_header
                                        ws.cell(row=1, column=1).value = report_client.upper()

                                        # header
                                        ws.cell(row=1, column=1).font = Font(
                                            name=header_font_name,
                                            size=header_font_size,
                                            bold=header_bold,
                                            italic=header_italic,
                                            color=header_font_color)                                     
                                        
                                        # subheader
                                        ws.cell(row=2, column=1).font = Font(
                                            name=subheader_font_name,
                                            size=subheader_font_size,
                                            bold=subheader_bold,
                                            italic=subheader_italic,
                                            color=subheader_font_color)
                                        
                                        ws.cell(row=5, column=1).font = Font(
                                            name='Arial',
                                            size=11,
                                            bold=True,
                                            color="0066CC")
                                        


                                        st.toast(f'Created {category_name} Sheet - Single')
                                        
                                    elif options=='Multiple':
                                        for company, data_frame in category_df.items():
                                            company = re.sub('/', '', company)
                                            data_frame = data_frame.rename(columns={'Bucket':'Company', 'Raw Date':'Date', 'AVE':'PR Value', 'Article ID':'Count'})
                                            data_frame.to_excel(writer, sheet_name=company, startrow=6, startcol=0, index=False)

                                            ws = writer.sheets[company]
                                            ws.cell(row=5, column=1).value = company                                        

                                        st.toast(f'Created {category_name} Sheet - Multiple')
                                
                                if cb_tonality_sheets:
                                    tone_data_frame_set = tonality_sheets(
                                        data_frame=df,
                                        category=main_category_select)
                                    
                                    for tone, data_frame_set in tone_data_frame_set.items():
                                        data_frame_set.to_excel(writer, sheet_name=tone, startrow=6, startcol=0, index=False)

                                        ws = writer.sheets[tone]
                                        ws.cell(row=5, column=1).value = tone
                                        ws.cell(row=2, column=1).value = sub_header
                                        ws.cell(row=1, column=1).value = report_client.upper()
                                
                                if daily_stats:
                                    daily_stats_dict = daily_statistics(
                                        data_frame=df,
                                        category=main_category_select)
                                    
                                    start_row = 0
                                    for _type, data_frame in daily_stats_dict.items():
                                        data_frame.to_excel(writer, sheet_name='Daily Statistics', startrow=start_row, startcol=16, index=False)
                                        start_row+=34
                                        
                                if media_stats:
                                    # st.write('Media Statistics')
                                    data_frame_dict = media_statistics(
                                        data_frame=df,
                                        category=main_category_select)
                                    
                                    start_row = 1
                                    for _type, data_frame in data_frame_dict.items():

                                        data_frame.to_excel(writer, sheet_name='Media Statistics', startrow=start_row, startcol=10, index=False)
                                        
                                        ws = writer.sheets['Media Statistics']
                                        if data_frame.shape[0]==10:
                                            ws.cell(row=start_row, column=11).value = f'Top 10 - {_type}'
                                        else:
                                            ws.cell(row=start_row, column=11).value = _type

                                        start_row+=25
                                
                                if share_voice:
                                    # st.write('Share of Voice')
                                    data_frame_set = share_of_voice(
                                        data_frame=df,
                                        category=main_category_select + competitor_category_select)
                                    
                                    start_row = 0
                                    start_col = 13
                                    table_count = 0
                                    for data_frame in data_frame_set:
                                        data_frame.to_excel(writer, sheet_name='Share of Voice', startrow=start_row, startcol=start_col, index=False)
                                        table_count+=1
                                        if table_count==2:
                                            start_col = 8
                                            
                                        start_row+=33
                                        
                                writer.close()
                                wb = load_workbook(f'{out_fname}.xlsx')
                                ws = wb['Sheet']
                                wb.remove(ws)
                                wb.save(f'{out_fname}.xlsx')
                                wb.close()

                                
                                # download area
                                result_file = open(report_name, 'rb')
                                st.success(f':red[NOTE:] Downloaded file will go to the :red[Downloads Folder]')
                                st.download_button(label='📥 Download Excel File', data=result_file, file_name=f'{out_fname} - Monthly_Statistics.xlsx')
                                    
                                    
